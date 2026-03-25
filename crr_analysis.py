"""
CRR Auction Analysis Tool
=========================
Reconstructs CRR contract value analysis from ERCOT auction result files.
Equivalent to the export5580 "Column1/Column2" computed columns and Sheet1 lookup table.

Usage:
    python crr_analysis.py

Output:
    crr_analysis_output.xlsx  - Excel file with:
        Sheet1: Hours lookup table (WDPEAK / WEPEAK / OFFPEAK hours for the target month)
        Data:   All contracts filtered by load zone, with hours and value columns
        Summary: Total pot by peak type, and customer proportion estimate
"""

import pandas as pd
import calendar
import zipfile
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIG — edit these to change the analysis
# ─────────────────────────────────────────────
TARGET_MONTH = 2        # February
TARGET_YEAR  = 2026
LOAD_ZONE    = "LZ_NORTH"
LOAD_ZONE_DIRECTION = "Sink"   # "Sink" = contracts delivering TO this zone

# Set customer avg load (MW) to calculate their proportional share.
# Set to None to skip that calculation.
CUSTOMER_AVG_LOAD_MW = None     # e.g. 150.0

# Total average load in the load zone (MW) — used as denominator for customer proportion.
# Get this from ERCOT's settlement/load data for the target month.
# Set to None to skip the proportion calculation.
TOTAL_ZONE_AVG_LOAD_MW = None   # e.g. 12500.0

# Files — can be a .csv path or a .zip path containing a MarketResults CSV.
# List annual sequences Seq6→Seq1 (oldest first), then the monthly.
# Each sequence is a separate auction; no CRR_ID overlap between sequences.
DATA_FILES = [
    # Annual 2026 1st6 — all 6 sequences (Seq6 = earliest Oct 2023, Seq1 = latest Nov 2025)
    "rpt.00011203.0000000000000000.20231005.100104738.20261st6AnnualAuctionSeq6CRRAuctionResults.zip",
    "rpt.00011203.0000000000000000.20240229.100116393.20261st6AnnualAuctionSeq5CRRAuctionResults.zip",
    "rpt.00011203.0000000000000000.20240801.080131669.20261st6AnnualAuctionSeq4CRRAuctionResults.zip",
    "rpt.00011203.0000000000000000.20250103.080110013.20261st6AnnualAuctionSeq3CRRAuctionResults.zip",
    "rpt.00011203.0000000000000000.20250605.080118891.20261st6AnnualAuctionSeq2CRRAuctionResults.zip",
    "rpt.00011203.0000000000000000.20251106.080135490.20261st6AnnualAuctionSeq1CRRAuctionResults (1).zip",
    # Monthly auction for the target month
    "rpt.00011201.0000000000000000.20260122.080059537.FEB2026MonthlyCRRAuctionResults.zip",
]

# For CRR type filtering — set to None to include all
INCLUDE_CRR_TYPES = None    # e.g. ["PREAWARD"] or ["STANDARD"] or None for all
# ─────────────────────────────────────────────


def _ercot_nerc_holidays(year: int) -> set:
    """
    Return ERCOT NERC holidays for a given year as a set of date objects.

    ERCOT uses 6 NERC holidays:
        New Year's Day     — Jan 1
        Memorial Day       — last Monday of May
        Independence Day   — Jul 4
        Labor Day          — first Monday of September
        Thanksgiving       — 4th Thursday of November
        Christmas          — Dec 25

    When a holiday falls on a Saturday → observed Friday
    When a holiday falls on a Sunday   → observed Monday
    Weekday holidays are counted as WEPEAK (not WDPEAK).
    """
    from datetime import date, timedelta

    def last_weekday_of_month(y, m, wd):
        """Last occurrence of weekday (Mon=0) in month."""
        d = date(y, m, calendar.monthrange(y, m)[1])
        while d.weekday() != wd:
            d -= timedelta(1)
        return d

    def nth_weekday_of_month(y, m, wd, n):
        """nth occurrence of weekday in month (1-indexed)."""
        d = date(y, m, 1)
        count = 0
        while True:
            if d.weekday() == wd:
                count += 1
                if count == n:
                    return d
            d += timedelta(1)

    def observe(d):
        if d.weekday() == 5:   # Saturday → Friday
            return d - timedelta(1)
        if d.weekday() == 6:   # Sunday → Monday
            return d + timedelta(1)
        return d

    from datetime import date
    holidays = {
        observe(date(year, 1,  1)),                          # New Year's Day
        last_weekday_of_month(year, 5, 0),                   # Memorial Day (last Mon May)
        observe(date(year, 7,  4)),                          # Independence Day
        nth_weekday_of_month(year, 9, 0, 1),                 # Labor Day (1st Mon Sep)
        nth_weekday_of_month(year, 11, 3, 4),                # Thanksgiving (4th Thu Nov)
        observe(date(year, 12, 25)),                         # Christmas
    }
    return holidays


def get_peak_hours_for_period(start_date, end_date) -> dict:
    """
    Return ERCOT peak hours for an arbitrary date range [start_date, end_date) exclusive.

    ERCOT peak definition:
        PeakWD   = HE07–HE22 on Mon–Fri EXCEPT NERC holidays → 16 hrs/day
        PeakWE   = HE07–HE22 on Sat–Sun + NERC holidays that fall on weekdays → 16 hrs/day
        Off-peak = all remaining hours

    start_date, end_date: datetime or date objects; end_date is EXCLUSIVE.
    """
    from datetime import date, timedelta

    if hasattr(start_date, 'date'):
        start_date = start_date.date()
    if hasattr(end_date, 'date'):
        end_date = end_date.date()

    # Build holiday sets for every year spanned
    years = range(start_date.year, end_date.year + 1)
    holidays = set()
    for y in years:
        holidays.update(_ercot_nerc_holidays(y))

    wdpeak = wepeak = offpeak = 0
    total = 0
    d = start_date
    while d < end_date:
        is_weekend = d.weekday() >= 5   # Sat or Sun
        is_holiday = d in holidays
        day_hours  = 24
        total     += day_hours

        if is_weekend or is_holiday:
            # Peak hours on this day count as WEPEAK
            wepeak  += 16
            offpeak += 8
        else:
            wdpeak  += 16
            offpeak += 8
        d += timedelta(1)

    return {
        "PeakWD":   wdpeak,
        "PeakWE":   wepeak,
        "Off-peak": offpeak,
        "Total":    total,
    }


def get_peak_hours(year: int, month: int) -> dict:
    """Return ERCOT peak hours for a single calendar month."""
    from datetime import date
    days = calendar.monthrange(year, month)[1]
    return get_peak_hours_for_period(
        date(year, month, 1),
        date(year, month, days) + __import__('datetime').timedelta(1),
    )


def load_market_results_csv(path: str) -> pd.DataFrame:
    """Read a MarketResults CSV, returning a raw DataFrame."""
    df = pd.read_csv(path)
    df["StartDate"] = pd.to_datetime(df["StartDate"])
    df["EndDate"]   = pd.to_datetime(df["EndDate"])
    df["_source_file"] = Path(path).name
    return df


def load_market_results(file_path: str) -> pd.DataFrame:
    """
    Load a MarketResults CSV from either a plain CSV or a ZIP archive.
    Automatically picks the MarketResults CSV inside a ZIP.
    """
    p = Path(file_path)
    if not p.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    if p.suffix.lower() == ".zip":
        with zipfile.ZipFile(p) as zf:
            market_files = [
                n for n in zf.namelist()
                if "MarketResults" in n and n.endswith(".csv")
            ]
            if not market_files:
                raise ValueError(f"No MarketResults CSV found in {file_path}")
            # If multiple, pick the one with no month sub-filter, or first one
            chosen = market_files[0]
            with zf.open(chosen) as f:
                df = pd.read_csv(f)
                df["StartDate"] = pd.to_datetime(df["StartDate"])
                df["EndDate"]   = pd.to_datetime(df["EndDate"])
                df["_source_file"] = Path(chosen).name
                return df
    else:
        return load_market_results_csv(file_path)


def filter_for_month(df: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    """Keep only contracts active during the target month."""
    days = calendar.monthrange(year, month)[1]
    month_start = pd.Timestamp(year, month, 1)
    month_end   = pd.Timestamp(year, month, days)
    mask = (df["StartDate"] <= month_end) & (df["EndDate"] >= month_start)
    return df[mask].copy()


def add_value_columns(df: pd.DataFrame, target_year: int, target_month: int) -> pd.DataFrame:
    """
    Add Column1 (hours) and Column2 ($ value) — mirrors the export5580 formula.

    Column1 = ERCOT peak hours over the CONTRACT'S full period (StartDate→EndDate).
              For monthly contracts this equals the single-month hours.
              For multi-month annual contracts this is the full period hours,
              which is what the MCP × hours × MW formula needs to produce
              the total contract value (as shown in the export5580 Cost$ column).

    Column2 = ShadowPricePerMWH × Column1 × MW
    """
    from datetime import date, timedelta

    peak_map = {"PeakWD": "PeakWD", "PeakWE": "PeakWE", "Off-peak": "Off-peak"}

    hours_cache = {}

    def row_hours(row):
        start = row["StartDate"]
        end   = row["EndDate"]
        if hasattr(start, 'date'):
            start = start.date()
        if hasattr(end, 'date'):
            end = end.date()
        # EndDate in ERCOT data is the last day of the contract (inclusive)
        end_excl = end + timedelta(1)
        key = (start, end_excl, row["TimeOfUse"])
        if key not in hours_cache:
            h = get_peak_hours_for_period(start, end_excl)
            hours_cache[key] = h
        return hours_cache[key].get(row["TimeOfUse"], 0)

    df = df.copy()
    df["Column1_hours"] = df.apply(row_hours, axis=1)
    df["Column2_value"] = df["ShadowPricePerMWH"] * df["Column1_hours"] * df["MW"]
    return df


def build_hours_lookup(year: int, month: int) -> pd.DataFrame:
    """
    Build the Sheet1 lookup table covering all contract period types (M through M6).

    Column1 hours differ by both peak type AND contract duration:
        M / M1  — contracts spanning only the target month
        M2–M6   — contracts spanning multiple months of the annual period
    """
    from datetime import date, timedelta

    month_name = datetime(year, month, 1).strftime("%B %Y")
    days = calendar.monthrange(year, month)[1]

    # For the annual 1st-half period, start = Jan 1, end varies
    # M  = just the target month
    # M1 = just the target month (same as M for monthly auction)
    # M2 = 2 months: Jan + target (or target + next, depending on context)
    # For 2026 1st6, the period is Jan–Jun and contracts start from Jan 1
    # We define M2..M6 as Jan 1 → end of month N
    annual_start = date(year, 1, 1)
    month_periods = {
        "M / M1":  (date(year, month, 1),        date(year, month, days)),
        "M2":      (annual_start,                 date(year, 2, calendar.monthrange(year, 2)[1])),
        "M3":      (annual_start,                 date(year, 3, calendar.monthrange(year, 3)[1])),
        "M4":      (annual_start,                 date(year, 4, calendar.monthrange(year, 4)[1])),
        "M5":      (annual_start,                 date(year, 5, calendar.monthrange(year, 5)[1])),
        "M6":      (annual_start,                 date(year, 6, calendar.monthrange(year, 6)[1])),
    }

    rows = []
    for label, (s, e) in month_periods.items():
        h = get_peak_hours_for_period(s, e + timedelta(1))
        rows.append({
            "Contract Period": label,
            "Period": f"{s.strftime('%b %d')}–{e.strftime('%b %d, %Y')}",
            "PeakWD Hours": h["PeakWD"],
            "PeakWE Hours": h["PeakWE"],
            "Off-peak Hours": h["Off-peak"],
            "Total Hours": h["Total"],
        })

    return pd.DataFrame(rows)


def run_analysis():
    print(f"\n{'='*60}")
    print(f"  CRR Analysis — {datetime(TARGET_YEAR, TARGET_MONTH, 1).strftime('%B %Y')}")
    print(f"  Load Zone : {LOAD_ZONE}  ({LOAD_ZONE_DIRECTION})")
    print(f"{'='*60}\n")

    # ── Hours lookup ──────────────────────────────────────────
    hours = get_peak_hours(TARGET_YEAR, TARGET_MONTH)
    hours_df = build_hours_lookup(TARGET_YEAR, TARGET_MONTH)
    print("Hours lookup (Sheet1):")
    print(hours_df.to_string(index=False))
    print()

    # ── Load & combine all data files ─────────────────────────
    frames = []
    for fpath in DATA_FILES:
        print(f"Loading: {fpath}")
        raw = load_market_results(fpath)
        filtered = filter_for_month(raw, TARGET_YEAR, TARGET_MONTH)
        frames.append(filtered)
        print(f"  → {len(raw):,} total rows  |  {len(filtered):,} active in "
              f"{datetime(TARGET_YEAR, TARGET_MONTH, 1).strftime('%b %Y')}")

    all_data = pd.concat(frames, ignore_index=True)

    # Remove duplicates (a contract may appear in both annual and monthly files)
    before_dedup = len(all_data)
    all_data = all_data.drop_duplicates(subset=["CRR_ID", "TimeOfUse"], keep="last")
    print(f"\nCombined: {before_dedup:,} rows → {len(all_data):,} after dedup\n")

    # ── Apply optional CRRType filter ─────────────────────────
    if INCLUDE_CRR_TYPES:
        all_data = all_data[all_data["CRRType"].isin(INCLUDE_CRR_TYPES)]
        print(f"Filtered to CRRType {INCLUDE_CRR_TYPES}: {len(all_data):,} rows\n")

    # ── Filter for load zone ──────────────────────────────────
    zone_data = all_data[all_data[LOAD_ZONE_DIRECTION] == LOAD_ZONE].copy()
    print(f"Contracts with {LOAD_ZONE_DIRECTION} = {LOAD_ZONE}: {len(zone_data):,}\n")

    # ── Add Column1 / Column2 ─────────────────────────────────
    print("Calculating contract period hours (Column1)...")
    zone_data = add_value_columns(zone_data, TARGET_YEAR, TARGET_MONTH)

    # ── Summary ───────────────────────────────────────────────
    total_value = zone_data["Column2_value"].sum()
    total_mw    = zone_data["MW"].sum()

    print("Contract value by Peak Type:")
    summary_rows = []
    for pt in ["PeakWD", "PeakWE", "Off-peak"]:
        sub = zone_data[zone_data["TimeOfUse"] == pt]
        val = sub["Column2_value"].sum()
        mw  = sub["MW"].sum()
        cnt = len(sub)
        print(f"  {pt:<12}  {cnt:>5} contracts  {mw:>10,.1f} MW  ${val:>14,.2f}")
        summary_rows.append({
            "Peak Type": pt,
            "Hours": hours[pt],
            "Contracts": cnt,
            "Total MW": round(mw, 2),
            "Total Value ($)": round(val, 2),
        })

    print(f"  {'─'*60}")
    print(f"  {'TOTAL':<12}  {len(zone_data):>5} contracts  "
          f"{total_mw:>10,.1f} MW  ${total_value:>14,.2f}\n")

    summary_rows.append({
        "Peak Type": "TOTAL",
        "Hours": hours["Total"],
        "Contracts": len(zone_data),
        "Total MW": round(total_mw, 2),
        "Total Value ($)": round(total_value, 2),
    })
    summary_df = pd.DataFrame(summary_rows)

    # ── Customer proportion (optional) ────────────────────────
    # Denominator: use TOTAL_ZONE_AVG_LOAD_MW if provided (from ERCOT load data),
    # otherwise fall back to total MW in CRR contracts as a rough proxy.
    if CUSTOMER_AVG_LOAD_MW is not None:
        denom = TOTAL_ZONE_AVG_LOAD_MW if TOTAL_ZONE_AVG_LOAD_MW else total_mw
        denom_label = "ERCOT zone avg load" if TOTAL_ZONE_AVG_LOAD_MW else "total CRR MW (proxy)"
        proportion = CUSTOMER_AVG_LOAD_MW / denom if denom else 0
        customer_pot = total_value * proportion
        print(f"Customer avg load        : {CUSTOMER_AVG_LOAD_MW:,.1f} MW")
        print(f"Zone denominator ({denom_label}): {denom:,.1f} MW")
        print(f"Customer proportion      : {proportion:.4%}")
        print(f"Customer potential CRR $ : ${customer_pot:,.2f}\n")

    # ── Write Excel output ────────────────────────────────────
    output_path = "crr_analysis_output.xlsx"
    write_excel(
        output_path=output_path,
        hours_df=hours_df,
        data_df=zone_data,
        summary_df=summary_df,
        year=TARGET_YEAR,
        month=TARGET_MONTH,
        load_zone=LOAD_ZONE,
    )
    print(f"Output written → {output_path}")
    return zone_data, summary_df


# ──────────────────────────────────────────────────────────────
# Excel writer
# ──────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
BOLD         = Font(bold=True)
CENTER       = Alignment(horizontal="center", vertical="center")
THIN         = Side(style="thin")
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def write_excel(output_path, hours_df, data_df, summary_df, year, month, load_zone):
    wb = openpyxl.Workbook()

    # ── Sheet1: Hours Lookup ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Sheet1"
    month_name = datetime(year, month, 1).strftime("%B %Y")

    ncols = len(hours_df.columns)
    ws1["A1"] = f"Contract Period Hours Lookup — {month_name}"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells(f"A1:{get_column_letter(ncols)}1")

    headers = list(hours_df.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for ri, row in enumerate(hours_df.itertuples(index=False), start=3):
        for ci, val in enumerate(row, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER

    _auto_width(ws1)

    # ── Summary sheet ─────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = f"CRR Value Summary — {load_zone} {LOAD_ZONE_DIRECTION} — {month_name}"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.merge_cells("A1:E1")

    sum_headers = list(summary_df.columns)
    for ci, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for ri, row in enumerate(summary_df.itertuples(index=False), start=3):
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER
            if row[0] == "TOTAL":
                cell.font = BOLD
            # Format dollar column
            if sum_headers[ci - 1] == "Total Value ($)":
                cell.number_format = "#,##0.00"
            if sum_headers[ci - 1] == "Total MW":
                cell.number_format = "#,##0.0"

    _auto_width(ws2)

    # ── Data sheet ────────────────────────────────────────────
    ws3 = wb.create_sheet("Data")

    # Reorder: put the key columns first
    display_cols = [
        "CRR_ID", "AccountHolder", "HedgeType", "BidType", "CRRType",
        "Source", "Sink", "StartDate", "EndDate", "TimeOfUse",
        "MW", "ShadowPricePerMWH", "Column1_hours", "Column2_value",
        "_source_file",
    ]
    # Only include columns that exist
    display_cols = [c for c in display_cols if c in data_df.columns]
    out_df = data_df[display_cols].copy()

    # Rename for clarity
    out_df = out_df.rename(columns={
        "Column1_hours":     "Column1 (Hours)",
        "Column2_value":     "Column2 (Value $)",
        "_source_file":      "Source File",
        "ShadowPricePerMWH": "MCP ($/MWh)",
    })

    for ci, h in enumerate(out_df.columns, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.fill = SUBHDR_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for ri, row in enumerate(out_df.itertuples(index=False), start=2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            if isinstance(val, pd.Timestamp):
                val = val.strftime("%m/%d/%Y")
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            col_name = out_df.columns[ci - 1]
            if col_name == "Column2 (Value $)":
                cell.number_format = "#,##0.00"
            elif col_name == "MCP ($/MWh)":
                cell.number_format = "0.000000"
            elif col_name == "Column1 (Hours)":
                cell.number_format = "0"

    _auto_width(ws3)
    ws3.freeze_panes = "A2"

    wb.save(output_path)


if __name__ == "__main__":
    run_analysis()
